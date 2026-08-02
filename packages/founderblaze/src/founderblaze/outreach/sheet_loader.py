from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

SUPPORTED = {".xlsx", ".xls", ".csv", ".xlsm"}
MAX_CHARS_PER_SHEET = 25_000
MAX_CHARS_TOTAL = 80_000

_REVENUE_HINTS = (
    "arr",
    "mrr",
    "revenue",
    "booking",
    "churn",
    "customer",
    "pipeline",
    "sales",
    "finance",
)


def load_workbook(file_path: str | Path) -> dict[str, Any]:
    resolved = Path(file_path).resolve()
    if not resolved.is_file():
        raise RuntimeError(f"Sheet file not found: {resolved}")
    ext = resolved.suffix.lower()
    if ext not in SUPPORTED:
        raise RuntimeError(
            f'Unsupported sheet type "{ext}". Use one of: {", ".join(sorted(SUPPORTED))}'
        )

    if ext == ".csv":
        sheets = [_load_csv(resolved)]
    else:
        sheets = _load_excel(resolved)

    if not sheets:
        raise RuntimeError("Workbook has no non-empty sheets")
    sheets.sort(key=lambda s: s["revenueScore"], reverse=True)
    return {
        "fileName": resolved.name,
        "sheets": sheets,
        "text": _build_combined_text(sheets),
    }


def _load_csv(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    headers = rows[0] if rows else []
    body = "\n".join(",".join(r) for r in rows)
    sample = body[:4000].lower()
    return {
        "sheetName": path.stem,
        "text": body[:MAX_CHARS_PER_SHEET],
        "rowCount": max(0, len(rows) - 1),
        "headers": headers,
        "revenueScore": _score(path.stem.lower()) * 3 + _score(sample),
    }


def _load_excel(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook as ox_load

    wb = ox_load(path, data_only=True, read_only=True)
    out: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(c) if c is not None else "" for c in first]
        lines = [",".join(headers)]
        count = 0
        for row in rows_iter:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            count += 1
            lines.append(",".join("" if c is None else str(c) for c in row))
        if count == 0 and not any(headers):
            continue
        csv_text = "\n".join(lines)[:MAX_CHARS_PER_SHEET]
        sample = csv_text[:4000].lower()
        out.append(
            {
                "sheetName": name,
                "text": csv_text,
                "rowCount": count,
                "headers": headers,
                "revenueScore": _score(name.lower()) * 3 + _score(sample),
            }
        )
    wb.close()
    return out


def _build_combined_text(sheets: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    used = 0
    for sheet in sheets:
        block = (
            f"===== SHEET: {sheet['sheetName']} =====\n"
            f"Rows: {sheet['rowCount']}\n"
            f"Columns: {', '.join(sheet['headers']) or '(none)'}\n"
            f"CSV:\n{sheet['text']}"
        )
        if used + len(block) > MAX_CHARS_TOTAL:
            remaining = max(0, MAX_CHARS_TOTAL - used - 80)
            if remaining > 500:
                parts.append(block[:remaining] + "\n…[truncated]")
            break
        parts.append(block)
        used += len(block)
    return "\n\n".join(parts)


def _score(text: str) -> int:
    return sum(1 for h in _REVENUE_HINTS if h in text)
